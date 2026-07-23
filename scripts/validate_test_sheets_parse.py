#!/usr/bin/env python3
"""Mirror of Code.gs ТЕСТ 1/2 parser — offline check against local xlsx dump."""

from __future__ import annotations

import calendar
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

ROOT = Path(__file__).resolve().parents[1]
CANDIDATES = [
    Path.home() / "AppData/Local/Temp/gsheets/data.xlsx",
    ROOT / ".tmp" / "original_client.xlsx",
]

SUPPLY_SHORT = {
    "Инженерное оборудование": "Инж. оборуд.",
    "Технологическое оборудование": "Тех. оборуд.",
    "Материалы": "Материалы",
    "Итого": "Итого",
}
FIN_MAP = {
    "Стоимость договора": "Договор · млн",
    "Стоимость по лимитам": "Лимиты · млн",
    "Оплачено подрядчику": "Оплачено · млн",
}
TASK_TITLE_HINTS = [
    ("рентабельности и прибыли", "Рентабельность АО «Успех»"),
    ("стратегию", "Стратегия ЖК «Северный»"),
    ("ндс на 2", "НДС +2% · разногласия"),
    ("дополнительных соглашений", "Доп. соглашения (НДС)"),
    ("сокращению расходов", "Сокращение расходов"),
    ("благоустройству", "ЖК «Солнце» · благоустр."),
    ("1с", "Контракты 1С"),
    ("материалов, находящихся на балансе", "Материалы на балансе"),
    ("субподряд", "Договоры с субподрядчиками"),
    ("папки с ответами", "Папки с ответами · ВПР"),
    ("претензионной", "Претензионная работа"),
    ("приказ об усилении", "Приказ · ответственность"),
    ("прут", "АО «Прут» · письмо"),
    ("берег", "ЖК «Берег» · справка"),
]


def cell(row, idx: int) -> str:
    if idx >= len(row) or row[idx] is None:
        return ""
    v = row[idx]
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.isoformat()
    return str(v).replace("\xa0", " ").replace("\r", "").strip()


def num(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("\xa0", "").replace(" ", "").replace("%", "").replace(",", ".")
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    return float(m.group(0)) if m else 0.0


def pct(v) -> float:
    n = num(v)
    if 0 < n <= 1.5:
        return round(n * 10000) / 100
    return n


def iso_clamp(y, m, d) -> str:
    if not y or not m or m < 1 or m > 12:
        return ""
    last = calendar.monthrange(y, m)[1]
    d = max(1, min(d, last))
    return f"{y:04d}-{m:02d}-{d:02d}"


def normalize_date(v) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip().replace("\xa0", " ").replace(" 00:00:00", "")
    s = re.sub(r"\.{2,}", ".", s)
    m = re.match(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})$", s)
    if m:
        return iso_clamp(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return s[:10]
    return ""


def is_section_header(a: str) -> bool:
    a = a.strip()
    if not a or len(a) > 120:
        return False
    if re.match(r"^\d+(\.\d+)?\.?\s+[А-ЯA-Z]", a) and re.search(r"[А-ЯA-Z]{4,}", a):
        if "%" in a or a.lower().startswith("готовность"):
            return False
        return True
    return False


def load(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        sheets[name] = [list(r) for r in wb[name].iter_rows(values_only=True)]
    return sheets


def find_sheet(sheets, want: str):
    for n in sheets:
        if n.strip().lower() == want.strip().lower():
            return sheets[n]
    raise KeyError(want)


def parse_general(rows):
    out = {}
    for r in rows:
        a, b = cell(r, 0), cell(r, 1)
        al = a.lower()
        if "наименование объекта" in al:
            out["object_name"] = b
        elif al.startswith("заказчик"):
            out["customer"] = b
        elif "генеральный подрядчик" in al:
            out["gen_contractor"] = b
        elif "стоимость работ" in al:
            out["contract_mln"] = num(b)
        elif "стоимость по лимитам" in al and "показател" not in al:
            out["limit_mln"] = num(b)
        elif "дата начала" in al:
            out["start_date"] = normalize_date(b) or normalize_date(r[1] if len(r) > 1 else "")
            if not out["start_date"] and isinstance(r[1], (date, datetime)):
                out["start_date"] = normalize_date(r[1])
        elif "контрактный срок" in al:
            out["contract_end"] = normalize_date(r[1] if len(r) > 1 else b)
        elif "директивный срок" in al:
            out["directive_end"] = normalize_date(b)
    return out


def parse_smr(rows):
    out = {"plan": None, "fact": None}
    in_sec = False
    for r in rows:
        a = cell(r, 0)
        al = a.lower()
        if "строительно-монтажные" in al:
            in_sec = True
            continue
        if in_sec and is_section_header(a) and "строительн" not in al:
            break
        if not in_sec:
            continue
        if "строительная готовность" in al:
            out["plan"] = pct(cell(r, 1) or (r[1] if len(r) > 1 else 0))
            out["fact"] = pct(cell(r, 2) or (r[2] if len(r) > 2 else 0))
            # raw numbers from openpyxl
            if isinstance(r[1], (int, float)):
                out["plan"] = pct(r[1])
            if isinstance(r[2], (int, float)):
                out["fact"] = pct(r[2])
    return out


def parse_supply(rows):
    items, total, in_sec = [], 0.0, False
    for r in rows:
        a = cell(r, 0)
        al = a.lower()
        if "комплектация" in al:
            in_sec = True
            continue
        if in_sec and is_section_header(a):
            break
        if not in_sec or "категория" in al or not a:
            continue
        raw_pct = r[3] if len(r) > 3 else 0
        p = pct(raw_pct)
        name = SUPPLY_SHORT.get(a, a)
        if a == "Итого":
            total = p
        tone = "alert" if p <= 25 else ("ok" if p >= 50 and name != "Итого" else "default")
        items.append({"name": name, "pct": p, "tone": tone})
    return items, total


def parse_finance(rows):
    items, paid, in_sec = [], 0, False
    for r in rows:
        a = cell(r, 0)
        al = a.lower()
        if "финансирование" in al or "взаиморасчёт" in al or "взаиморасчет" in al:
            in_sec = True
            continue
        if in_sec and is_section_header(a):
            break
        if not in_sec or "показател" in al or not a:
            continue
        label = FIN_MAP.get(a)
        if not label:
            continue
        plan = round(num(r[1] if len(r) > 1 else 0))
        fact = round(num(r[2] if len(r) > 2 else 0))
        items.append({"name": label, "plan": plan, "fact": fact})
        if label == "Оплачено · млн":
            paid = fact
    return items, paid


def parse_labor(rows):
    labor, in_sec = [], False
    for r in rows:
        a = cell(r, 0)
        al = a.lower()
        if "трудовые ресурсы" in al:
            in_sec = True
            continue
        if in_sec and is_section_header(a):
            break
        if not in_sec:
            continue
        if "рабочие" in al:
            labor.append({"name": "Рабочие · чел.", "plan": round(num(r[1])), "fact": round(num(r[2]))})
        elif "итр" in al:
            labor.append({"name": "ИТР · чел.", "plan": round(num(r[1])), "fact": round(num(r[2]))})
    return labor


def parse_risks(rows):
    risks, in_sec = [], False
    cur_p = cur_i = cur_l = ""
    header = False
    for r in rows:
        a = cell(r, 0)
        al = a.lower()
        if "проблем" in al and ("риск" in al or "решен" in al):
            in_sec = True
            continue
        if not in_sec:
            continue
        if a in ("№", "No") or (al.startswith("проблема") and cell(r, 1)):
            header = True
            continue
        if not header:
            continue
        problem, impact, level = cell(r, 1), cell(r, 2), cell(r, 3)
        solution, owner = cell(r, 4), cell(r, 5)
        deadline_raw, status = cell(r, 6), cell(r, 7)
        # raw dates
        if len(r) > 6 and isinstance(r[6], (date, datetime)):
            deadline_raw = normalize_date(r[6])
        is_new = bool(re.match(r"^\d+\.?$", a.replace(" ", "")))
        if is_new and problem:
            cur_p, cur_i, cur_l = problem, impact, level
        sol = solution
        if not sol and is_new and status:
            sol = cur_p[:120]
        if not sol and not status and not deadline_raw:
            continue
        if not cur_p and not problem:
            continue
        dl = normalize_date(deadline_raw) if not re.match(r"^\d{4}-", deadline_raw or "") else deadline_raw
        if not dl:
            dl = normalize_date(deadline_raw)
        risks.append(
            {
                "problem": cur_p or problem,
                "solution": sol,
                "owner": owner,
                "deadline": dl,
                "status": status or "В работе",
            }
        )
    return risks


def is_bullet(s: str) -> bool:
    s = s.strip()
    return bool(s) and s[0] in "-–—"


def strip_bullet(s: str) -> str:
    return re.sub(r"^[\s\-–—]+", "", s).rstrip(";").strip()


def parse_cadence(raw: str) -> str:
    s = (raw or "").lower().replace("\n", " ")
    if "постоян" in s:
        return "постоянно"
    if "еженед" in s or "еженедель" in s:
        if "пятниц" in s or re.search(r"\bпт\b", s):
            return "еженедельно (пт)"
        if "понедельник" in s or re.search(r"\bпн\b", s):
            return "еженедельно (пн)"
        return "еженедельно"
    return ""


def parse_tasks(rows):
    tasks = []
    block, owner = "", ""
    start = 1 if rows and "задача" in cell(rows[0], 1).lower() else 0
    for r in rows[start:]:
        c0, c1, c2, c3, c4 = cell(r, 0), cell(r, 1), cell(r, 2), cell(r, 3), cell(r, 4)
        if re.match(r"^план на неделю", c0, re.I) or re.match(r"^план на неделю", c1, re.I):
            if block and "План на неделю" not in block:
                block = block + " / План на неделю"
            continue
        if re.match(r"^задачи:?$", c0, re.I) or re.match(r"^задачи:?$", c1, re.I):
            continue
        if re.match(r"^\d+(\.\d+)?\.?$", c0.replace(" ", "")) and c1 and not is_bullet(c1):
            block = re.sub(r"\s*/\s*Задачи:?\s*$", "", c1, flags=re.I).strip()
            owner = c2
            continue
        if c0 and not is_bullet(c0) and re.search(r"ФИНАНС|ЗАКУПК|ПРАВОВ|КАЗНАЧ|КАЦ", c0, re.I):
            block = c0
            owner = c1
            continue
        if is_bullet(c1) or (c1.startswith("-") if c1 else False):
            text, resp, draw, st = strip_bullet(c1), c2 or owner, c3, c4
        elif is_bullet(c0) or (c0.startswith("-") if c0 else False):
            text, resp, draw, st = strip_bullet(c0), c1 or owner, c2, c3
        else:
            continue
        if isinstance(r[3] if len(r) > 3 else None, (date, datetime)) and is_bullet(c1):
            draw = normalize_date(r[3])
        if isinstance(r[2] if len(r) > 2 else None, (date, datetime)) and is_bullet(c0):
            draw = normalize_date(r[2])
        cadence = parse_cadence(draw)
        deadline = "" if cadence else normalize_date(draw)
        dtype = cadence or ("дата" if deadline else "")
        if not st:
            st = "В работе"
        title = next((t for k, t in TASK_TITLE_HINTS if k in text.lower()), text[:42])
        tasks.append({"t": title, "description": text, "status": st, "block": block, "deadline": deadline, "deadline_type": dtype, "responsible": resp})
    return tasks


def main() -> int:
    path = next((p for p in CANDIDATES if p.exists()), None)
    if not path:
        print("No xlsx found")
        return 1
    sheets = load(path)
    t1 = find_sheet(sheets, "ТЕСТ 1")
    t2 = find_sheet(sheets, "ТЕСТ 2")
    g = parse_general(t1)
    smr = parse_smr(t1)
    supply, supply_total = parse_supply(t1)
    finance, paid = parse_finance(t1)
    labor = parse_labor(t1)
    risks = parse_risks(t1)
    tasks = parse_tasks(t2)

    # Prefer raw openpyxl for readiness if display strings odd
    print("object:", g.get("object_name"))
    print("gen:", g.get("gen_contractor"))
    print("start/end:", g.get("start_date"), g.get("directive_end"))
    print("smr:", smr)
    print("supply_total:", supply_total, "n=", len(supply))
    print("finance:", finance, "paid", paid)
    print("labor:", labor)
    print("risks:", len(risks))
    for x in risks:
        print(" ", x["status"], x["deadline"], (x["solution"] or "")[:50])
    print("tasks:", len(tasks))
    for t in tasks:
        print(" ", t["block"][:20], "|", t["t"], "|", t["status"], t.get("deadline_type") or t.get("deadline"))

    ok = True
    if g.get("object_name") != "ЖК «Северный»":
        print("FAIL object_name")
        ok = False
    if smr.get("fact") is None or abs(smr["fact"] - 54.71) > 0.05:
        print("FAIL smr fact", smr)
        ok = False
    if abs(supply_total - 33) > 0.05:
        print("FAIL supply total", supply_total)
        ok = False
    if paid != 4368:
        print("FAIL paid", paid)
        ok = False
    if len(tasks) < 12:
        print("FAIL tasks count", len(tasks))
        ok = False
    if len(risks) < 5:
        print("FAIL risks count", len(risks))
        ok = False
    out = ROOT / ".tmp" / "parsed_from_test_sheets.json"
    out.parent.mkdir(exist_ok=True)
    out.write_text(
        json.dumps(
            {"project": g, "smr": smr, "supply": supply, "finance": finance, "labor": labor, "risks": risks, "tasks": tasks},
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print("wrote", out)
    print("OK" if ok else "ISSUES")
    return 0 if ok else 2


if __name__ == "__main__":
    raise SystemExit(main())
