# -*- coding: utf-8 -*-
"""Build Looker-ready xlsx from data/*.csv"""
from __future__ import annotations

import csv
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
OUT = DATA / "Dashboard_Mobile_Source.xlsx"

# CSV file -> sheet name for Looker / Google Sheets
SHEETS = [
    ("01_project_card.csv", "project_card"),
    ("02_kpi_summary.csv", "kpi_summary"),
    ("03_smr.csv", "smr"),
    ("04_labor.csv", "labor"),
    ("05_supply.csv", "supply"),
    ("06_finance.csv", "finance"),
    ("07_risks.csv", "risks"),
    ("08_tasks.csv", "tasks"),
]

HEADER_FILL = PatternFill("solid", fgColor="F7F4EF")
HEADER_FONT = Font(name="Calibri", bold=True, color="1A1A1A", size=11)
CELL_FONT = Font(name="Calibri", color="1A1A1A", size=11)
THIN = Border(
    bottom=Side(style="thin", color="E4DFD7"),
)


def read_csv(path: Path) -> list[list[str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.reader(f))


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    for csv_name, sheet_name in SHEETS:
        path = DATA / csv_name
        if not path.exists():
            raise SystemExit(f"Missing {path}")
        rows = read_csv(path)
        ws = wb.create_sheet(sheet_name)
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(r_idx, c_idx, value)
                cell.font = HEADER_FONT if r_idx == 1 else CELL_FONT
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if r_idx == 1:
                    cell.fill = HEADER_FILL
                else:
                    cell.border = THIN
        # Autosize-ish widths
        for col in ws.columns:
            letter = col[0].column_letter
            maxlen = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[letter].width = min(max(maxlen + 2, 12), 48)
        ws.freeze_panes = "A2"

    # README sheet for the client / assembler
    info = wb.create_sheet("README", 0)
    lines = [
        ["Dashboard_Mobile_Source — данные для Looker Studio"],
        [""],
        ["Как использовать"],
        ["1. Загрузить этот файл в Google Диск → Открыть с помощью → Google Таблицы"],
        ["2. В Looker Studio: Создать → Отчёт → Google Таблицы → эта книга"],
        ["3. Подключать листы: project_card, risks, tasks, finance (остальные — по необходимости)"],
        [""],
        ["Стиль: ПРИРОДНЫЙ МИНИМАЛИЗМ"],
        ["Фон #F7F5F0 | карточки #FFFFFF | текст #3A3A3A / #8A8580"],
        ["Акценты #2C5F4A #B8864E #9E3B33 #6B8F7A | линии #E8E0D8"],
        ["Инструкция: docs/STYLE.md + docs/INSTRUCTIONS.md"],
        [""],
        ["Не подключать сырые ТЕСТ 1 / ТЕСТ 2 как основной источник"],
    ]
    for r, row in enumerate(lines, start=1):
        info.cell(r, 1, row[0]).font = Font(name="Calibri", size=12, bold=(r == 1))
    info.column_dimensions["A"].width = 100

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
